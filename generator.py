import argparse
from datetime import datetime
import openpyxl
import random
import re
import string


# Create an argument parser.
parser = argparse.ArgumentParser(description='Generate random test data into SQL insert statements using formatted strings in an Excel file.')
# Add arguments.
parser.add_argument('-n', '--rows', metavar='rows', type=int, default=10, required=False, help='the number of insert statements to generate (default = 10)')
# Parse the arguments.
args = parser.parse_args()

# Assume we have something like AX5[A-Z]{3}[0-9]{4}[A-Z]{2}[0-9]{1}

# Load the first and last names.
first_names_txt = 'first-names.txt'
last_names_txt = 'last-names.txt'
with open(first_names_txt) as file:
    first_names = [name.strip() for name in file.read().split('\n')]
with open(last_names_txt) as file:
    last_names = [name.strip() for name in file.read().split('\n')]

# Substitute the patterns.
# Uppercase letters.
def replace_lowercase(pattern: str) -> str:
    for sub_pattern in re.findall('\[a\-z\]\{\d+\}', pattern):
        length = int(re.findall('(?<=\{)\d+(?=\})', sub_pattern)[0])
        replacement = ''.join(random.choice(string.ascii_lowercase) for _ in range(length))
        pattern = pattern.replace(sub_pattern, replacement, 1)
    return pattern


# Lowercase letters.
def replace_uppercase(pattern: str) -> str:
    """Replace uppercase placeholder values with uppercase letters"""
    for sub_pattern in re.findall('\[A\-Z\]\{\d+\}', pattern):
        length = int(re.findall('(?<=\{)\d+(?=\})', sub_pattern)[0])
        replacement = ''.join(random.choice(string.ascii_uppercase) for _ in range(length))
        pattern = pattern.replace(sub_pattern, replacement, 1)
    return pattern


# Digits
def replace_digits(pattern: str) -> str:
    """Replace integer placeholder values with integers"""
    for sub_pattern in re.findall('\[\d+\-\d+\]\{\d+\}', pattern):
        length = int(re.findall('(?<=\{)\d+(?=\})', sub_pattern)[0])
        start, end = [int(d) for d in re.findall('\d+\-\d+', sub_pattern)[0].split('-')]
        replacement = ''.join(str(random.randint(start, end)) for _ in range(length))
        pattern = pattern.replace(sub_pattern, replacement, 1)
    return pattern


# Decimals.
def replace_decimals(pattern: str) -> str:
    """Replace decimal placeholder values with decimals"""
    for sub_pattern in re.findall('\[\d+\.\d+\-\d+\.\d+\]', pattern):
        start, end = re.findall('\d+\.\d+', sub_pattern)
        decimal_places = len(start.split('.')[1])
        start = float(start)*pow(10, decimal_places)
        end = float(end)*pow(10, decimal_places)
        replacement = str(random.randrange(start, end)/pow(10, decimal_places))
        pattern = pattern.replace(sub_pattern, replacement, 1)
    return pattern

# Alphanumeric
def replace_alphanum(pattern: str) -> str:
    """Replace alphanumeric placeholder values with alphanumeric values"""
    for sub_pattern in re.findall('\{\{alphanum\}\}\{\d+\}', pattern):
        length = int(re.findall('\d+', sub_pattern)[0])
        replacement = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(length))
        pattern = pattern.replace(sub_pattern, replacement)
    return pattern


# Lists of values
def replace_lists(pattern: str) -> str:
    """Replace list placeholder values with lists of values"""
    for sub_pattern in re.findall('\[[\w\d\s\,]+\]\{\d+\}', pattern):
        length = int(re.findall('(?<=\{)\d+(?=\})', sub_pattern)[0])
        value_list = [value.strip() for value in re.findall('(?<=\[)[\w\d\s\,]+(?=\])', sub_pattern)[0].split(',')]
        replacement = ''.join(random.choice(value_list) for _ in range(length))
        pattern = pattern.replace(sub_pattern, replacement)
    return pattern


# Replace names.
def replace_names(pattern: str, first_name: str, last_name) -> str:
    """Replace name placeholder values with first and last names"""
    email = first_name + random.choice(['.', '_', '']) + last_name + '@' + random.choice(['example.com', 'example.net', 'test.com', 'test.net', 'mail.com', 'mail.net', 'gmail.com', 'yahoo.com', 'outlook.com'])
    primary_address = ''.join(random.choice(string.digits) for _ in range(random.randint(2, 4))) + ' ' + random.choice(last_names).upper() + ' ' + random.choice(['RD', 'LN', 'AVE', 'BLVD', 'ST', 'PL', 'WAY'])
    secondary_address = random.choice(['APT ', 'STE' '#']) + ''.join(random.choice(string.digits) for _ in range(random.randint(1, 4)))
    pattern = pattern.replace('{{first_name}}', first_name)
    pattern = pattern.replace('{{last_name}}', last_name)
    pattern = pattern.replace('{{email}}', email)
    pattern = pattern.replace('{{primary_address}}', primary_address)
    pattern = pattern.replace('{{primary_address}}', secondary_address)
    return pattern

# Replace dates.
def replace_dates(pattern: str) -> str:
    for sub_pattern in re.findall(r'{{[MDY\-\/]+}}', pattern):
        replacement = re.findall(r'(?<={{)[MDY\-\/]+(?=}})', sub_pattern)[0]
        month = str(random.randint(1, 12))
        day = str(random.randint(1, 28 if month == 2 else 30 if month in [4, 6, 9] else 31))
        current_year = datetime.now().year
        year = str(random.randint(current_year - 20, current_year))

        # Replace the month
        replacement = replacement.replace('MM', f'{month:0>2}')       # Zero-padded
        replacement = replacement.replace('M', month)                 # Non-zer-padded
        # Replace the day
        replacement = replacement.replace('DD', f'{day:0>2}')         # Zero-padded
        replacement = replacement.replace('D', day)                   # Non-zero-padded
        # Replace the year
        replacement = replacement.replace('YYYY', year)               # Full year format
        replacement = replacement.replace('YY', year[:2])             # Half year format

        # Replace the sub-pattern with the replacement
        pattern = pattern.replace(sub_pattern, replacement, 1)

    return pattern


##### Loading the Excel file #####
excel_file = 'tables.xlsx'
# load the Excel file into a workbook object.
print(f'[*] Loading {excel_file}.')
wb = openpyxl.load_workbook(excel_file)
print('[*] Done.')
# Initialize a dictionary of tables.
tables = {sheetname: {} for sheetname in wb.sheetnames}
# Initialize a SQL string.
sql = ''
n = args.rows
print('[*] Performing inserts.')
for i in range(n):
    print(f'[*] On row {i + 1} of {n}...\r', end='')
    sql += f'-- Insert {i+1} of {n}\n'
    # Initialize a dictionary of aliases.
    aliases = {}
    # Generate a random first and last name.
    first_name = random.choice(first_names)
    last_name = random.choice(last_names)
    # Iterate through each worksheet and populate the table dictionary.
    for table in list(tables):
        # Get the worksheet.
        ws = wb[table]
        # Create a table dictionary in the tables dictionary.
        tables[table] = {}
        # Iterate through the worksheet.
        for row in ws.iter_rows():
            attribute = row[0].value
            pattern = row[1].value
            alias = row[2].value if len(row) == 3 else ''
            # Make pattern substitutions.
            pattern = replace_lowercase(pattern)
            pattern = replace_uppercase(pattern)
            pattern = replace_digits(pattern)
            pattern = replace_decimals(pattern)
            pattern = replace_alphanum(pattern)
            pattern = replace_dates(pattern)
            pattern = replace_lists(pattern)
            pattern = replace_names(pattern, first_name, last_name)
            # Add the alias if one is set.
            if alias and alias not in aliases:
                aliases[alias] = pattern
            # Create an attribute dictionary in the table dictinoary.
            tables[table][attribute] = {}
            # Add the values to the the dictionary.
            tables[table][attribute]['pattern'] = pattern
            tables[table][attribute]['alias'] = alias
        # Generate the SQL.
        fields = list(tables[table])
        values = [tables[table][field]['pattern'] if not tables[table][field]['alias'] else aliases[tables[table][field]['alias']] for field in fields]
        sql += f'INSERT INTO {table} '
        sql += '(' + ', '.join(fields) + ') '
        sql += 'VALUES (' + ', '.join([f"'{value}'" for value in values]) + ')'
        sql += '\n'
    sql += '\n'
print('\n[*] Done.')
print(f'[*] Closing {excel_file}.')
# Close the workbook.
wb.close()
print('[*] Done.')
# Write the SQL to a file.
sql_file = 'inserts.sql'
print(f'[*] Writing to {sql_file}')
with open(sql_file, 'w') as file:
    file.write(sql)
print('[*] Done.')
print('[*] Exiting.')

exit(0)

